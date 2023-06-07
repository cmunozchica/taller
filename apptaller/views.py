from django.shortcuts import render, redirect
from rest_framework.response import Response
from .models import *
from django.core.paginator import Paginator
from django.http import Http404
from django.db.models import Q
from .utils import render_to_pdf
from django.http import HttpResponse
from django.views.generic import View
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .forms import EquipoForm, AlistamientoForm, UserRegisterForm, DiagnosticoForm2, InsumosTecnicoForm, HerramientaForm
from django.contrib import messages
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils.six import BytesIO
from django.conf import settings
from qrcode import *
import time

# Create your views here.
def index(request):
    
    return render(request, 'apptaller/base.html')

    


def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data['username']
            messages.success(request, f'Usuario {username} creado')
            return redirect('index')
    else:
        form = UserRegisterForm()
    context = { 'form' : form }
    return render(request, 'register.html', context) 

# Vistas de clientes

def clienteList(request):
    busqueda = request.GET.get("buscar")
    cliente_queryset = Cliente.objects.all()
    page = request.GET.get('page', 1)

    if busqueda:
            cliente_queryset = Cliente.objects.filter(
                Q(id__icontains = busqueda) |
                Q(razon_social__icontains = busqueda)             
            ).distinct()
    try:        
        paginator = Paginator(cliente_queryset,10)    
        cliente_queryset = paginator.page(page)
    except:
        raise Http404  

    context = {'cliente' : cliente_queryset, 'paginator': paginator}
    return render(request, 'apptaller/clientes-list.html', context)

def addCliente(request):
   if request.method == "POST":
       nit = request.POST.get('nit')       
       razon_social = request.POST.get('razon_social')
       direccion = request.POST.get('direccion')
       telefono = request.POST.get('telefono')
       email = request.POST.get('email')
       ciudad = request.POST.get('ciudad')
       contacto = request.POST.get('contacto')

       cliente = Cliente(
           nit = nit,
           razon_social = razon_social,
           direccion = direccion,
           telefono = telefono,
           email = email,
           ciudad = ciudad,
           contacto = contacto,
       )
       cliente.save()
       return redirect('cliente-list')

   return render(request, 'cliente-list.html') 

def editCliente(request):
        cliente = Cliente.objects.all()

        context = {'cliente' : cliente,}
        return render(request, 'cliente-list.html') 

def updateCliente(request, id):
    if request.method == "POST":
       nit = request.POST.get('nit')       
       razon_social = request.POST.get('razon_social')
       direccion = request.POST.get('direccion')
       telefono = request.POST.get('telefono')
       email = request.POST.get('email')
       ciudad = request.POST.get('ciudad')
       contacto = request.POST.get('contacto')

       cliente = Cliente(
           id = id,
           nit = nit,
           razon_social = razon_social,
           direccion = direccion,
           telefono = telefono,
           email = email,
           ciudad = ciudad,
           contacto = contacto,
       )
       cliente.save()
       return redirect('cliente-list')

    return render(request, 'cliente-list.html') 

def deleteCliente(request,id):
    cliente = Cliente.objects.filter(id = id).delete()
    context = {'cliente' : cliente,}
    return redirect('cliente-list')

class ListClientesPdf(View):

    def get(self, request, *args, **kwargs):
        cliente_queryset = Cliente.objects.all()
        context = {
            'cliente_context': cliente_queryset, 'count': cliente_queryset.count(),
        }
        pdf = render_to_pdf('apptaller/reporte-cliente-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')

class reporteClienteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        clientes =  Cliente.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(
            start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True)
        ws['B1'] = 'REPORTE DE CLIENTES'
        ws.merge_cells('B1:H1')
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True)
        ws['B3'] = 'NIT'
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True)
        ws['C3'] = 'RAZON SOCIAL'
        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True)
        ws['D3'] = 'DIRECCION'
        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True)
        ws['E3'] = 'TELEFONO'
        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True)
        ws['F3'] = 'EMAIL'
        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True)
        ws['G3'] = 'CIUDAD'
        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True)
        ws['H3'] = 'CONTACTO'

        cont = 7

        for cliente in clientes:
            ws.cell(row=cont, column=2).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=2).value = cliente.nit
            ws.cell(row=cont, column=3).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=3).value = cliente.razon_social
            ws.cell(row=cont, column=4).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=4).value = cliente.direccion
            ws.cell(row=cont, column=5).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=5).value = cliente.telefono
            ws.cell(row=cont, column=6).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=6).value = cliente.email
            ws.cell(row=cont, column=7).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=7).value = cliente.ciudad
            ws.cell(row=cont, column=8).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=8).value = cliente.contacto
            cont += 1

        nombre_archivo = "ReporteTecnicosExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
# Vistas de tecnicos

def tecnicoList(request):
    busqueda = request.GET.get("buscar")
    tecnico_queryset = Tecnico.objects.all()
    page = request.GET.get('page', 1)

    if busqueda:
            tecnico_queryset = Tecnico.objects.filter(
                Q(id__icontains = busqueda) |
                Q(cedula__icontains = busqueda) |
                Q(nombre__icontains = busqueda) |
                Q(apellido__icontains = busqueda)             
            ).distinct()
    try:        
        paginator = Paginator(tecnico_queryset,10)    
        tecnico_queryset = paginator.page(page)
    except:
        raise Http404  

    context = {'tecnico' : tecnico_queryset, 'paginator': paginator}
    return render(request, 'apptaller/tecnico-list.html', context)

def addTecnico(request):
   if request.method == "POST":
       cedula = request.POST.get('cedula')       
       nombre = request.POST.get('nombre')
       apellido = request.POST.get('apellido')
       cargo = request.POST.get('cargo')
       celular = request.POST.get('celular')
       correo = request.POST.get('correo')
       estado = request.POST.get('estado')

       tecnico = Tecnico(
           cedula = cedula,
           nombre = nombre,
           apellido = apellido,
           cargo = cargo,
           celular = celular,
           correo = correo,
           estado = estado,
       )
       tecnico.save()
       return redirect('tecnico-list')

   return render(request, 'tecnico-list.html') 

def editTecnico(request):
        tecnico = Tecnico.objects.all()

        context = {'tecnico' : tecnico,}
        return render(request, 'tecnico-list.html') 

def updateTecnico(request, id):
    if request.method == "POST":
       cedula = request.POST.get('cedula')       
       nombre = request.POST.get('nombre')
       apellido = request.POST.get('apellido')
       cargo = request.POST.get('cargo')
       celular = request.POST.get('celular')
       correo = request.POST.get('correo')
       estado = request.POST.get('estado')
       created = request.POST.get('created')

       tecnico = Tecnico(
           id = id,
           cedula = cedula,
           nombre = nombre,
           apellido = apellido,
           cargo = cargo,
           celular = celular,
           correo = correo,
           estado = estado,
           created = created,
       )
       tecnico.save()
       return redirect('tecnico-list')

    return render(request, 'tecnico-list.html') 

def deleteTecnico(request,id):
    tecnico = Tecnico.objects.filter(id = id).delete()
    context = {'tecnico' : tecnico,}
    return redirect('tecnico-list')

class ListTecnicosPdf(View):

    def get(self, request, *args, **kwargs):
        tecnico_queryset = Tecnico.objects.all()
        context = {
            'tecnico_context': tecnico_queryset, 'count': tecnico_queryset.count(),
        }
        pdf = render_to_pdf('apptaller/reporte-tecnico-pdf.html', context)
        return HttpResponse(pdf, content_type='application/pdf')

class reporteTecnicoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tecnicos =  Tecnico.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(
            start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True)
        ws['B1'] = 'REPORTE DE TECNICOS'
        ws.merge_cells('B1:H1')
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True)
        ws['B3'] = 'CEDULA'
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True)
        ws['C3'] = 'NOMBRE'
        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True)
        ws['D3'] = 'APELLIDO'
        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True)
        ws['E3'] = 'CARGO'
        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True)
        ws['F3'] = 'CELULAR'
        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True)
        ws['G3'] = 'CORREO'
        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(
            start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True)
        ws['H3'] = 'ESTADO'

        cont = 7

        for i in tecnicos:
            ws.cell(row=cont, column=2).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=2).value = i.cedula
            ws.cell(row=cont, column=3).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=3).value = i.nombre
            ws.cell(row=cont, column=4).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=4).value = i.apellido
            ws.cell(row=cont, column=5).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=5).value = i.cargo
            ws.cell(row=cont, column=6).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=6).value = i.celular
            ws.cell(row=cont, column=7).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=7).value = i.correo
            ws.cell(row=cont, column=8).alignment = Alignment(
                horizontal="left")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=8).value = i.estado
            cont += 1

        nombre_archivo = "ReporteTecnicoExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
#vistas marca
def marcaList(request):

    marca_queryset = Marca.objects.all()
 

    context = {'marca' : marca_queryset}

def addMarca(request):
   if request.method == "POST":
       marca = request.POST.get('marca')       


       marca = Marca(
           marca = marca,

       )
       marca.save()
       return redirect('index')


   return render(request, 'apptaller/base.html') 

def modeloList(request):

    modelo_queryset = Modelo.objects.all()
 

    context = {'marca' : modelo_queryset}

def addModelo(request):
   if request.method == "POST":
       modelo = request.POST.get('modelo')       


       modelo = Modelo(
           modelo = modelo,

       )
       modelo.save()
       return redirect('index')

   return render(request, 'apptaller/base.html') 

def equipoList(request):
    busqueda = request.GET.get("buscar")
    equipo_queryset = Equipo.objects.all().order_by('id')
    page = request.GET.get('page', 1)

    if busqueda:
            equipo_queryset = Equipo.objects.filter(
                Q(id__icontains = busqueda) |
                Q(serie__icontains = busqueda)
            ).distinct()
    try:        
        paginator = Paginator(equipo_queryset,20)    
        equipo_queryset = paginator.page(page)
    except:
        raise Http404    

    context = {'equipo':equipo_queryset, 'paginator': paginator}
    return render(request, "apptaller/equipo-list.html", context)

def addEquipo(request):
    if request.method == "POST":  
        form = EquipoForm(request.POST)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('equipo-list')  
            except:  
                pass  
    else:  
        form = EquipoForm()  
    return render(request,'apptaller/add-equipo.html',{'form':form}) 




def updateEquipo(request, id):
    equipo = Equipo.objects.get(id=id)
    form = EquipoForm(initial={'serie': equipo.serie, 'marca': equipo.marca, 'modelo': equipo.modelo, 'accesorios': equipo.accesorio, 'contador': equipo.contador, "update": equipo.update})
    if request.method == "POST":  
        form = EquipoForm(request.POST, instance=equipo)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('/equipo-list/')  
            except Exception as e: 
                pass    
    return render(request,'apptaller/equipoUpdate.html',{'form':form}) 

def deleteEquipo(request,id):
    equipo = Equipo.objects.filter(id = id).delete()
    context = {'equipo' : equipo,}
    return redirect('equipo-list')

def uploadFileServ(request):
    if request.method == "POST":
        # Fetching the form data
        fileTitle = request.POST["fileTitle"]
        uploadedFile = request.FILES["uploadedFile"]

        # Saving the information in the database
        document = ManualServicio(
            title = fileTitle,
            uploadedFile = uploadedFile
        )
        document.save()
        return redirect('manualServicio-list') 

    documents = ManualServicio.objects.all()

    return render(request, "apptaller/uploadServ.html", context = {
        "files": documents
    })

def manualServicio_list_view(request):
    busqueda = request.GET.get("buscar")
    manualServicio_queryset = ManualServicio.objects.all().order_by('id')


    if busqueda:
            manualServicio_queryset = ManualServicio.objects.filter(
                Q(id__icontains = busqueda) |                
                Q(title__icontains = busqueda) 
            ).distinct()
    paginator = Paginator(manualServicio_queryset,10)
    page = request.GET.get('page')
    manualServicio_queryset = paginator.get_page(page)

    context = {'manualServicio_context':manualServicio_queryset}
    return render(request, "apptaller/listarManualSer.html", context)

def deleteManualServ(request,id):
    manualServ = ManualServicio.objects.filter(id = id).delete()
    context = {'manualServ' : manualServ,}
    return redirect('manualServicio-list')

def uploadFilePart(request):
    if request.method == "POST":
        # Fetching the form data
        fileTitle = request.POST["fileTitle"]
        uploadedFile = request.FILES["uploadedFile"]

        # Saving the information in the database
        document = ManualParte(
            title = fileTitle,
            uploadedFile = uploadedFile
        )
        document.save()
        return redirect('manualParte-list') 

    documents = ManualParte.objects.all()

    return render(request, "apptaller/uploadPart.html", context = {
        "files": documents
    })

def manualParte_list_view(request):
    busqueda = request.GET.get("buscar")
    manualParte_queryset = ManualParte.objects.all().order_by('id')


    if busqueda:
            manualParte_queryset = ManualParte.objects.filter(
                Q(id__icontains = busqueda) |                
                Q(title__icontains = busqueda) 
            ).distinct()
    paginator = Paginator(manualParte_queryset,10)
    page = request.GET.get('page')
    manualParte_queryset = paginator.get_page(page)

    context = {'manualParte_context':manualParte_queryset}
    return render(request, "apptaller/listarManualPart.html", context)


def numParteList(request):
    busqueda = request.GET.get("buscar")
    numParte_queryset = NumParte.objects.all()
    page = request.GET.get('page', 1)

    if busqueda:
            numParte_queryset = NumParte.objects.filter(
                Q(id__icontains = busqueda) |
                Q(modelo__icontains = busqueda) |
                Q(numero_parte__icontains = busqueda) |
                Q(descripcion__icontains = busqueda)             
            ).distinct()
    try:        
        paginator = Paginator(numParte_queryset,10)    
        numParte_queryset = paginator.page(page)
    except:
        raise Http404  

    context = {'numparte' : numParte_queryset, 'paginator': paginator}
    return render(request, 'apptaller/numParte-list.html', context)

def addNumParte(request):
   if request.method == "POST":
       marca = request.POST.get('marca')       
       modelo = request.POST.get('modelo')
       numero_parte = request.POST.get('numero_parte')
       descripcion = request.POST.get('descripcion')


       numParte = NumParte(
           marca = marca,
           modelo = modelo,
           numero_parte = numero_parte,
           descripcion = descripcion,

       )
       numParte.save()
       return redirect('numParte-list')

   return render(request, 'numParte-list.html') 

def editNumParte(request):
        numParte = NumParte.objects.all()

        context = {'numParte' :  numParte,}
        return render(request, ' numParte-list.html') 

def updateNumParte(request, id):
    if request.method == "POST":
       marca = request.POST.get('marca')       
       modelo = request.POST.get('modelo')
       numero_parte = request.POST.get('numero_parte')
       descripcion = request.POST.get('descripcion')


       numParte = NumParte(
           id = id,
           marca = marca,
           modelo = modelo,
           numero_parte = numero_parte,
           descripcion = descripcion,

       )
       numParte.save()
       return redirect('numParte-list')

    return render(request, 'numParte-list.html') 

def deleteNumParte(request,id):
    numParte = NumParte.objects.filter(id = id).delete()
    context = {'numParte' : numParte,}
    return redirect('numParte-list')


def alistamientoList(request):
    busqueda = request.GET.get("buscar")
    alistamiento_queryset = Alistamiento.objects.all().order_by('id')
    page = request.GET.get('page', 1)

    if busqueda:
            alistamiento_queryset = Alistamiento.objects.filter(
                Q(id__icontains = busqueda) |
                Q(serie__icontains = busqueda) |
                Q(estado__icontains = busqueda) |
                Q(created__icontains = busqueda) |
                Q(update__icontains = busqueda)   
            ).distinct()
    try:        
        paginator = Paginator(alistamiento_queryset,8)    
        alistamiento_queryset = paginator.page(page)
    except:
        raise Http404    

    context = {'alistamiento':alistamiento_queryset, 'paginator': paginator}
    return render(request, "apptaller/alistamiento-list.html", context)


def addAlistamiento(request):
    if request.method == "POST":  
        form = AlistamientoForm(request.POST)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('alistamiento-list')  
            except:  
                pass  
    else:  
        form = AlistamientoForm()  
    return render(request,'apptaller/add-alistamiento.html',{'form':form}) 




def updateAlistamiento(request, id):
    alistamiento = Alistamiento.objects.get(id=id)
    form = AlistamientoForm(initial={'serie': alistamiento.serie, 'marca': alistamiento.marca, 'modelo': alistamiento.modelo, 'tecnico': alistamiento.tecnico, 'accesorios': alistamiento.accesorio, 'contador': alistamiento.contador, "update": alistamiento.update})
    if request.method == "POST":  
        form = AlistamientoForm(request.POST, instance=alistamiento)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('/alistamiento-list/')  
            except Exception as e: 
                pass    
    return render(request,'apptaller/alistamientoUpdate.html',{'form':form}) 

def deleteAlistamiento(request,id):
    alistamiento = Alistamiento.objects.filter(id = id).delete()
    context = {'alistamiento' : alistamiento,}
    return redirect('alistamiento-list')

def diagnosticoList(request):
    busqueda = request.GET.get("buscar")
    diagnostico_queryset = DiagnosticoEquipos.objects.all().order_by('id')
    page = request.GET.get('page', 1)

    if busqueda:
            diagnostico_queryset = DiagnosticoEquipos.objects.filter(
                Q(id__icontains = busqueda) |
                Q(serie__icontains = busqueda) |
                Q(modelo__icontains = busqueda) |
                Q(tecnico__icontains = busqueda) |
                Q(created__icontains = busqueda)
            ).distinct()
    try:        
        paginator = Paginator(diagnostico_queryset,15)    
        diagnostico_queryset = paginator.page(page)
    except:
        raise Http404    

    context = {'diagnostico':diagnostico_queryset, 'paginator': paginator}
    return render(request, "apptaller/diagnostico-list.html", context)

def addDiagnostico(request):
    if request.method == "POST":  
        form = DiagnosticoForm2(request.POST)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('diagnostico-list')  
            except:  
                pass  
    else:  
        form = DiagnosticoForm2()  
    return render(request,'apptaller/add-diagnostico.html',{'form':form}) 

def insumoTecnicoList(request):
    busqueda = request.GET.get("buscar")
    insumo_queryset = InsumosTecnicos.objects.all().order_by('id')
    page = request.GET.get('page', 1)

    if busqueda:
            insumo_queryset = InsumosTecnicos.objects.filter(
                Q(id__icontains = busqueda) |
                Q(insumo__icontains = busqueda) |
                Q(tecnico__icontains = busqueda)   
            ).distinct()
    try:        
        paginator = Paginator(insumo_queryset,8)    
        insumo_queryset = paginator.page(page)
    except:
        raise Http404    

    context = {'insumo':insumo_queryset, 'paginator': paginator}
    return render(request, "apptaller/insumo-list.html", context)

def addInsumo(request):
    if request.method == "POST":  
        form = InsumosTecnicoForm(request.POST)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('insumo-list')  
            except:  
                pass  
    else:  
        form = InsumosTecnicoForm()  
    return render(request,'apptaller/add-insumo.html',{'form':form}) 

def herramientaList(request):
    busqueda = request.GET.get("buscar")
    herramienta_queryset = HerramientaTecnicos.objects.all().order_by('id')
    page = request.GET.get('page', 1)

    if busqueda:
            herramienta_queryset = HerramientaTecnicos.objects.filter(
                Q(id__icontains = busqueda) |
                Q(herramienta__icontains = busqueda) |
                Q(prestamo__icontains = busqueda) |
                Q(tecnico__icontains = busqueda)   
            ).distinct()
    try:        
        paginator = Paginator(herramienta_queryset,8)    
        herramienta_queryset = paginator.page(page)
    except:
        raise Http404    

    context = {'insumo':herramienta_queryset, 'paginator': paginator}
    return render(request, "apptaller/herramienta-list.html", context)

def addHerramienta(request):
    if request.method == "POST":  
        form = HerramientaForm(request.POST)  
        if form.is_valid():  
            try:  
                form.save() 
                model = form.instance
                return redirect('herramienta-list')  
            except:  
                pass  
    else:  
        form = HerramientaForm()  
    return render(request,'apptaller/add-herramienta.html',{'form':form}) 



def qr_gen(request):

    if request.method == 'POST':
        data = request.POST['data']
        img = make(data)
        img_name="qr" + str(time.time()) + '.png'
        img.save(settings.MEDIA_ROOT + '/' + img_name)
        return render(request, 'apptaller/qrcode.html', {'img_name': img_name})
    return render(request, 'apptaller/qrcode.html')
